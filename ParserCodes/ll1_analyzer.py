"""
LL(1) Parser Analyzer
Checks if a grammar is LL(1) by computing FIRST and FOLLOW sets
and building the predictive parsing table
"""

from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from grammar import grammar



class LL1Analyzer:
    def __init__(self, grammar_text):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start_symbol = None
        
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        self.parsing_table = defaultdict(dict)
        self.conflicts = []
        
        self.parse_grammar(grammar_text)
        self.compute_first()
        self.compute_follow()
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if self.start_symbol is None:
                    self.start_symbol = lhs
                
                if lhs not in self.grammar:
                    self.grammar[lhs] = []
                self.grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol != 'ε' and symbol not in self.non_terminals:
                        self.terminals.add(symbol)
        
        self.terminals.add('$')  # End of input marker
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        if rhs == 'ε' or rhs == 'epsilon':
            return ['ε']
        
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols if symbols else ['ε']
    
    def compute_first(self):
        """Compute FIRST sets for all symbols"""
        print("\n=== COMPUTING FIRST SETS ===\n")
        
        # Initialize FIRST for terminals
        for t in self.terminals:
            self.first[t].add(t)
        
        self.first['ε'].add('ε')
        
        # Compute FIRST for non-terminals
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    before = len(self.first[A])
                    
                    # If production is ε
                    if production == ['ε']:
                        self.first[A].add('ε')
                        if len(self.first[A]) > before:
                            changed = True
                        continue
                    
                    # For each symbol in production
                    all_nullable = True
                    for symbol in production:
                        # Add FIRST(symbol) - {ε} to FIRST(A)
                        self.first[A].update(self.first.get(symbol, set()) - {'ε'})
                        
                        # If symbol doesn't have ε in its FIRST set, stop
                        if 'ε' not in self.first.get(symbol, set()):
                            all_nullable = False
                            break
                    
                    # If all symbols are nullable, add ε to FIRST(A)
                    if all_nullable:
                        self.first[A].add('ε')
                    
                    if len(self.first[A]) > before:
                        changed = True
        
        # Print FIRST sets
        for A in sorted(self.non_terminals):
            print(f"FIRST({A}) = {{{', '.join(sorted(self.first[A]))}}}")
    
    def compute_first_of_string(self, symbols):
        """Compute FIRST of a string of symbols"""
        result = set()
        
        if not symbols or symbols == ['ε']:
            result.add('ε')
            return result
        
        all_nullable = True
        for symbol in symbols:
            # Add FIRST(symbol) - {ε}
            result.update(self.first.get(symbol, set()) - {'ε'})
            
            # If symbol is not nullable, stop
            if 'ε' not in self.first.get(symbol, set()):
                all_nullable = False
                break
        
        # If all symbols are nullable, add ε
        if all_nullable:
            result.add('ε')
        
        return result
    
    def compute_follow(self):
        """Compute FOLLOW sets for all non-terminals"""
        print("\n=== COMPUTING FOLLOW SETS ===\n")
        
        # Add $ to FOLLOW of start symbol
        self.follow[self.start_symbol].add('$')
        
        changed = True
        while changed:
            changed = False
            
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    # Skip epsilon productions
                    if production == ['ε']:
                        continue
                    
                    # For each symbol in production
                    for i, B in enumerate(production):
                        if B in self.non_terminals:
                            before = len(self.follow[B])
                            
                            # Get symbols after B
                            beta = production[i+1:]
                            
                            # Add FIRST(β) - {ε} to FOLLOW(B)
                            first_beta = self.compute_first_of_string(beta)
                            self.follow[B].update(first_beta - {'ε'})
                            
                            # If β is nullable (or empty), add FOLLOW(A) to FOLLOW(B)
                            if not beta or 'ε' in first_beta:
                                self.follow[B].update(self.follow[A])
                            
                            if len(self.follow[B]) > before:
                                changed = True
        
        # Print FOLLOW sets
        for A in sorted(self.non_terminals):
            print(f"FOLLOW({A}) = {{{', '.join(sorted(self.follow[A]))}}}")
    
    def build_parsing_table(self):
        """Build LL(1) predictive parsing table"""
        print("\n=== BUILDING LL(1) PARSING TABLE ===\n")
        
        for A in self.non_terminals:
            for production in self.grammar.get(A, []):
                prod_num = self.get_production_number(A, production)
                prod_str = f"{A} → {' '.join(production)}"
                
                # Compute FIRST(α) where α is the production
                first_alpha = self.compute_first_of_string(production)
                
                # For each terminal in FIRST(α) - {ε}
                for terminal in first_alpha - {'ε'}:
                    if terminal in self.terminals:
                        if terminal in self.parsing_table[A]:
                            # Conflict detected
                            existing = self.parsing_table[A][terminal]
                            self.conflicts.append({
                                'non_terminal': A,
                                'terminal': terminal,
                                'production1': existing,
                                'production2': prod_str,
                                'type': 'FIRST-FIRST conflict'
                            })
                            self.parsing_table[A][terminal] = f"{existing} / {prod_str}"
                        else:
                            self.parsing_table[A][terminal] = prod_str
                
                # If ε ∈ FIRST(α), for each terminal in FOLLOW(A)
                if 'ε' in first_alpha:
                    for terminal in self.follow[A]:
                        if terminal in self.terminals:
                            if terminal in self.parsing_table[A]:
                                # Conflict detected
                                existing = self.parsing_table[A][terminal]
                                self.conflicts.append({
                                    'non_terminal': A,
                                    'terminal': terminal,
                                    'production1': existing,
                                    'production2': prod_str,
                                    'type': 'FIRST-FOLLOW conflict'
                                })
                                self.parsing_table[A][terminal] = f"{existing} / {prod_str}"
                            else:
                                self.parsing_table[A][terminal] = prod_str
        
        print(f"Parsing table entries: {sum(len(row) for row in self.parsing_table.values())}")
        
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} conflict(s)!")
        else:
            print("\n✓ No conflicts found - Grammar is LL(1)!")
    
    def get_production_number(self, lhs, rhs):
        """Get production number for A → α"""
        prod_num = 0
        for nt in sorted(self.grammar.keys()):
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def is_ll1(self):
        """Check if grammar is LL(1)"""
        return len(self.conflicts) == 0
    
    def check_ll1_conditions(self):
        """Check specific LL(1) conditions and return violations"""
        violations = []
        
        for A in self.non_terminals:
            productions = self.grammar.get(A, [])
            
            # Check for each pair of productions
            for i, prod1 in enumerate(productions):
                first1 = self.compute_first_of_string(prod1)
                
                for j, prod2 in enumerate(productions):
                    if i >= j:
                        continue
                    
                    first2 = self.compute_first_of_string(prod2)
                    
                    # Condition 1: FIRST sets must be disjoint
                    intersection = (first1 - {'ε'}) & (first2 - {'ε'})
                    if intersection:
                        violations.append({
                            'condition': 'FIRST sets not disjoint',
                            'non_terminal': A,
                            'production1': f"{A} → {' '.join(prod1)}",
                            'production2': f"{A} → {' '.join(prod2)}",
                            'intersection': intersection
                        })
                    
                    # Condition 2: At most one production can derive ε
                    if 'ε' in first1 and 'ε' in first2:
                        violations.append({
                            'condition': 'Multiple ε-productions',
                            'non_terminal': A,
                            'production1': f"{A} → {' '.join(prod1)}",
                            'production2': f"{A} → {' '.join(prod2)}",
                            'intersection': set()
                        })
                    
                    # Condition 3: If one can derive ε, FIRST and FOLLOW must be disjoint
                    if 'ε' in first1:
                        follow_first_intersection = (first2 - {'ε'}) & self.follow[A]
                        if follow_first_intersection:
                            violations.append({
                                'condition': 'FIRST-FOLLOW conflict',
                                'non_terminal': A,
                                'production1': f"{A} → {' '.join(prod1)} (nullable)",
                                'production2': f"{A} → {' '.join(prod2)}",
                                'intersection': follow_first_intersection
                            })
                    
                    if 'ε' in first2:
                        follow_first_intersection = (first1 - {'ε'}) & self.follow[A]
                        if follow_first_intersection:
                            violations.append({
                                'condition': 'FIRST-FOLLOW conflict',
                                'non_terminal': A,
                                'production1': f"{A} → {' '.join(prod1)}",
                                'production2': f"{A} → {' '.join(prod2)} (nullable)",
                                'intersection': follow_first_intersection
                            })
        
        return violations
    
    def generate_excel(self, filename='ll1_analysis.xlsx'):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_first_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_parsing_table_sheet(wb)
        self.create_conditions_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "LL(1) Parser Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Is LL(1)?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_ll1 = self.is_ll1()
        ws['B3'] = "YES ✓" if is_ll1 else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_ll1 else "FF0000")
        
        ws['A5'] = "Number of Non-Terminals:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.non_terminals)
        
        ws['A6'] = "Number of Terminals:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = len(self.terminals) - 1  # Exclude $
        
        ws['A7'] = "Number of Conflicts:"
        ws['A7'].font = Font(bold=True)
        ws['B7'] = len(self.conflicts)
        
        if self.conflicts:
            ws['A9'] = "Conflicts Details:"
            ws['A9'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A9:E9')
            
            ws['A11'] = "Non-Terminal"
            ws['B11'] = "Terminal"
            ws['C11'] = "Conflict Type"
            ws['D11'] = "Production 1"
            ws['E11'] = "Production 2"
            
            for cell in ['A11', 'B11', 'C11', 'D11', 'E11']:
                ws[cell].font = Font(bold=True, size=11)
                ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            
            row = 12
            for conflict in self.conflicts:
                ws[f'A{row}'] = conflict['non_terminal']
                ws[f'B{row}'] = conflict['terminal']
                ws[f'C{row}'] = conflict['type']
                ws[f'D{row}'] = conflict['production1']
                ws[f'E{row}'] = conflict['production2']
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        else:
            ws['A9'] = "✓ No conflicts found - Grammar is LL(1)!"
            ws['A9'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A9:D9')
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 35
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Production #"
        ws['B1'] = "Non-Terminal"
        ws['C1'] = "Production"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        prod_num = 0
        
        for nt in sorted(self.grammar.keys()):
            for prod in self.grammar[nt]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = nt
                ws[f'C{row}'] = f"{nt} → {' '.join(prod)}"
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def create_first_sheet(self, wb):
        """Create FIRST sets sheet"""
        ws = wb.create_sheet("FIRST Sets")
        
        ws['A1'] = "Symbol"
        ws['B1'] = "FIRST Set"
        ws['C1'] = "Explanation"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            first_set = sorted(self.first.get(A, set()))
            ws[f'B{row}'] = ', '.join(first_set)
            
            # Check if nullable
            if 'ε' in self.first.get(A, set()):
                ws[f'C{row}'] = "Nullable (can derive ε)"
                ws[f'C{row}'].font = Font(color="0000FF", italic=True)
            else:
                ws[f'C{row}'] = "Not nullable"
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
    
    def create_follow_sheet(self, wb):
        """Create FOLLOW sets sheet"""
        ws = wb.create_sheet("FOLLOW Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        ws['C1'] = "Explanation"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            follow_set = sorted(self.follow.get(A, set()))
            ws[f'B{row}'] = ', '.join(follow_set)
            
            if A == self.start_symbol:
                ws[f'C{row}'] = "Start symbol (contains $)"
                ws[f'C{row}'].font = Font(color="0000FF", italic=True)
            else:
                ws[f'C{row}'] = "Computed from grammar rules"
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 35
    
    def create_parsing_table_sheet(self, wb):
        """Create LL(1) parsing table sheet"""
        ws = wb.create_sheet("Parsing Table")
        
        terminals_sorted = sorted(self.terminals)
        
        ws['A1'] = "Non-Terminal"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        for nt in sorted(self.non_terminals):
            ws.cell(row, 1, nt)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in self.parsing_table[nt]:
                    entry = self.parsing_table[nt][terminal]
                    ws.cell(row, col_idx, entry)
                    ws.cell(row, col_idx).alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Highlight conflicts
                    if '/' in entry:
                        ws.cell(row, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row, col_idx).font = Font(color="FFFFFF", bold=True)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    def create_conditions_sheet(self, wb):
        """Create LL(1) conditions check sheet"""
        ws = wb.create_sheet("LL(1) Conditions")
        
        ws['A1'] = "LL(1) Conditions Verification"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A3'] = "For a grammar to be LL(1), for each non-terminal A with productions A → α₁ | α₂ | ... | αₙ:"
        ws.merge_cells('A3:E3')
        ws['A3'].alignment = Alignment(wrap_text=True)
        
        ws['A4'] = "1. FIRST(αᵢ) ∩ FIRST(αⱼ) = ∅ for all i ≠ j"
        ws['A5'] = "2. At most one αᵢ can derive ε"
        ws['A6'] = "3. If αᵢ ⇒* ε, then FIRST(αⱼ) ∩ FOLLOW(A) = ∅ for all i ≠ j"
        
        ws['A8'] = "Condition"
        ws['B8'] = "Non-Terminal"
        ws['C8'] = "Production 1"
        ws['D8'] = "Production 2"
        ws['E8'] = "Conflict Symbols"
        
        for cell in ['A8', 'B8', 'C8', 'D8', 'E8']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=11, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        violations = self.check_ll1_conditions()
        
        if not violations:
            ws['A10'] = "✓ All LL(1) conditions satisfied!"
            ws['A10'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A10:E10')
        else:
            row = 9
            for violation in violations:
                ws[f'A{row}'] = violation['condition']
                ws[f'B{row}'] = violation['non_terminal']
                ws[f'C{row}'] = violation['production1']
                ws[f'D{row}'] = violation['production2']
                ws[f'E{row}'] = ', '.join(sorted(violation['intersection'])) if violation['intersection'] else 'N/A'
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20


def main():
    # Example grammar - you can replace this with your grammar
#     grammar = """
# E → T E'
# E' → '+' T E' | ε
# T → F T'
# T' → '*' F T' | ε
# F → '(' E ')' | id
# """
    
    print("="*70)
    print("LL(1) PARSER ANALYZER")
    print("="*70)
    
    analyzer = LL1Analyzer(grammar)
    
    print(f"\nGrammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)-1} terminals")
    print(f"Start symbol: {analyzer.start_symbol}")
    
    analyzer.build_parsing_table()
    analyzer.generate_excel()
    
    print("\n" + "="*70)
    print("FINAL RESULT")
    print("="*70)
    
    if analyzer.is_ll1():
        print("✅ The grammar IS LL(1)")
        print("\n🎉 The grammar can be parsed using predictive parsing!")
    else:
        print("❌ The grammar IS NOT LL(1)")
        print(f"\nFound {len(analyzer.conflicts)} conflict(s):")
        for i, conflict in enumerate(analyzer.conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    Non-Terminal: {conflict['non_terminal']}")
            print(f"    Terminal: {conflict['terminal']}")
            print(f"    Type: {conflict['type']}")
            print(f"    Production 1: {conflict['production1']}")
            print(f"    Production 2: {conflict['production2']}")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()