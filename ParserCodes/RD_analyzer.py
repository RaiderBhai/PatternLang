"""
Recursive Descent Parser Analyzer
Checks if a grammar is suitable for Recursive Descent parsing
A grammar is suitable for Recursive Descent if:
1. It has no left recursion (direct or indirect)
2. It is LL(1) or can be made LL(1) through backtracking
"""

from collections import defaultdict, deque
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from grammar import grammar


class RecursiveDescentAnalyzer:
    def __init__(self, grammar_text):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start_symbol = None
        
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        self.parsing_table = defaultdict(dict)
        
        # Recursive Descent specific checks
        self.left_recursive_nts = []
        self.direct_left_recursion = []
        self.indirect_left_recursion = []
        self.left_factoring_needed = []
        self.conflicts = []
        
        self.parse_grammar(grammar_text)
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if self.start_symbol is None:
                    self.start_symbol = lhs
                
                if lhs not in self.grammar:
                    self.grammar[lhs] = []
                self.grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol != 'ε' and symbol not in self.non_terminals:
                        self.terminals.add(symbol)
        
        self.terminals.add('$')
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        if rhs == 'ε' or rhs == 'epsilon':
            return ['ε']
        
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols if symbols else ['ε']
    
    def check_direct_left_recursion(self):
        """Check for direct left recursion: A → A α"""
        print("\n=== CHECKING DIRECT LEFT RECURSION ===\n")
        
        for A in self.non_terminals:
            for production in self.grammar.get(A, []):
                if production and production[0] == A:
                    self.direct_left_recursion.append({
                        'non_terminal': A,
                        'production': f"{A} → {' '.join(production)}",
                        'type': 'Direct'
                    })
                    if A not in self.left_recursive_nts:
                        self.left_recursive_nts.append(A)
        
        if self.direct_left_recursion:
            print(f"⚠️  Found {len(self.direct_left_recursion)} direct left recursion(s):")
            for dlr in self.direct_left_recursion:
                print(f"  - {dlr['production']}")
        else:
            print("✓ No direct left recursion found")
    
    def check_indirect_left_recursion(self):
        """Check for indirect left recursion using graph traversal"""
        print("\n=== CHECKING INDIRECT LEFT RECURSION ===\n")
        
        # Build derivation graph
        derives_to = defaultdict(set)
        for A in self.non_terminals:
            for production in self.grammar.get(A, []):
                if production and production != ['ε'] and production[0] in self.non_terminals:
                    derives_to[A].add(production[0])
        
        # Check for cycles using DFS
        def has_cycle(node, visited, rec_stack, path):
            visited.add(node)
            rec_stack.add(node)
            path.append(node)
            
            for neighbor in derives_to.get(node, set()):
                if neighbor not in visited:
                    if has_cycle(neighbor, visited, rec_stack, path):
                        return True
                elif neighbor in rec_stack:
                    # Found a cycle
                    cycle_start = path.index(neighbor)
                    cycle = path[cycle_start:] + [neighbor]
                    
                    # Check if this is left recursion (not just any cycle)
                    if cycle[0] == cycle[-1]:
                        cycle_str = ' → '.join(cycle)
                        self.indirect_left_recursion.append({
                            'cycle': cycle,
                            'description': cycle_str,
                            'type': 'Indirect'
                        })
                        for nt in cycle[:-1]:
                            if nt not in self.left_recursive_nts:
                                self.left_recursive_nts.append(nt)
                    return True
            
            path.pop()
            rec_stack.remove(node)
            return False
        
        visited = set()
        for node in self.non_terminals:
            if node not in visited:
                has_cycle(node, visited, set(), [])
        
        if self.indirect_left_recursion:
            print(f"⚠️  Found {len(self.indirect_left_recursion)} indirect left recursion(s):")
            for ilr in self.indirect_left_recursion:
                print(f"  - {ilr['description']}")
        else:
            print("✓ No indirect left recursion found")
    
    def check_left_factoring(self):
        """Check if left factoring is needed"""
        print("\n=== CHECKING LEFT FACTORING ===\n")
        
        for A in self.non_terminals:
            productions = self.grammar.get(A, [])
            
            # Group productions by first symbol
            first_symbols = defaultdict(list)
            for prod in productions:
                if prod != ['ε']:
                    first_sym = prod[0]
                    first_symbols[first_sym].append(prod)
            
            # Check for common prefixes
            for first_sym, prods in first_symbols.items():
                if len(prods) > 1:
                    # Find longest common prefix
                    common_prefix = self.find_common_prefix(prods)
                    if len(common_prefix) > 0:
                        self.left_factoring_needed.append({
                            'non_terminal': A,
                            'common_prefix': common_prefix,
                            'productions': [f"{A} → {' '.join(p)}" for p in prods],
                            'first_symbol': first_sym
                        })
        
        if self.left_factoring_needed:
            print(f"⚠️  Left factoring needed for {len(self.left_factoring_needed)} case(s):")
            for lf in self.left_factoring_needed:
                print(f"  - {lf['non_terminal']}: common prefix = {' '.join(lf['common_prefix'])}")
        else:
            print("✓ No left factoring needed")
    
    def find_common_prefix(self, productions):
        """Find the longest common prefix among productions"""
        if not productions:
            return []
        
        min_len = min(len(p) for p in productions)
        common = []
        
        for i in range(min_len):
            symbol = productions[0][i]
            if all(p[i] == symbol for p in productions):
                common.append(symbol)
            else:
                break
        
        return common
    
    def compute_first(self):
        """Compute FIRST sets for all symbols"""
        print("\n=== COMPUTING FIRST SETS ===\n")
        
        # Initialize FIRST for terminals
        for t in self.terminals:
            self.first[t].add(t)
        
        self.first['ε'].add('ε')
        
        # Compute FIRST for non-terminals
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    before = len(self.first[A])
                    
                    if production == ['ε']:
                        self.first[A].add('ε')
                        if len(self.first[A]) > before:
                            changed = True
                        continue
                    
                    all_nullable = True
                    for symbol in production:
                        self.first[A].update(self.first.get(symbol, set()) - {'ε'})
                        
                        if 'ε' not in self.first.get(symbol, set()):
                            all_nullable = False
                            break
                    
                    if all_nullable:
                        self.first[A].add('ε')
                    
                    if len(self.first[A]) > before:
                        changed = True
        
        for A in sorted(self.non_terminals):
            print(f"FIRST({A}) = {{{', '.join(sorted(self.first[A]))}}}")
    
    def compute_first_of_string(self, symbols):
        """Compute FIRST of a string of symbols"""
        result = set()
        
        if not symbols or symbols == ['ε']:
            result.add('ε')
            return result
        
        all_nullable = True
        for symbol in symbols:
            result.update(self.first.get(symbol, set()) - {'ε'})
            
            if 'ε' not in self.first.get(symbol, set()):
                all_nullable = False
                break
        
        if all_nullable:
            result.add('ε')
        
        return result
    
    def compute_follow(self):
        """Compute FOLLOW sets for all non-terminals"""
        print("\n=== COMPUTING FOLLOW SETS ===\n")
        
        self.follow[self.start_symbol].add('$')
        
        changed = True
        while changed:
            changed = False
            
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    if production == ['ε']:
                        continue
                    
                    for i, B in enumerate(production):
                        if B in self.non_terminals:
                            before = len(self.follow[B])
                            
                            beta = production[i+1:]
                            first_beta = self.compute_first_of_string(beta)
                            self.follow[B].update(first_beta - {'ε'})
                            
                            if not beta or 'ε' in first_beta:
                                self.follow[B].update(self.follow[A])
                            
                            if len(self.follow[B]) > before:
                                changed = True
        
        for A in sorted(self.non_terminals):
            print(f"FOLLOW({A}) = {{{', '.join(sorted(self.follow[A]))}}}")
    
    def build_parsing_table(self):
        """Build LL(1) predictive parsing table"""
        print("\n=== BUILDING PARSING TABLE ===\n")
        
        for A in self.non_terminals:
            for production in self.grammar.get(A, []):
                prod_str = f"{A} → {' '.join(production)}"
                first_alpha = self.compute_first_of_string(production)
                
                for terminal in first_alpha - {'ε'}:
                    if terminal in self.terminals:
                        if terminal in self.parsing_table[A]:
                            existing = self.parsing_table[A][terminal]
                            self.conflicts.append({
                                'non_terminal': A,
                                'terminal': terminal,
                                'production1': existing,
                                'production2': prod_str,
                                'type': 'FIRST-FIRST conflict'
                            })
                            self.parsing_table[A][terminal] = f"{existing} / {prod_str}"
                        else:
                            self.parsing_table[A][terminal] = prod_str
                
                if 'ε' in first_alpha:
                    for terminal in self.follow[A]:
                        if terminal in self.terminals:
                            if terminal in self.parsing_table[A]:
                                existing = self.parsing_table[A][terminal]
                                self.conflicts.append({
                                    'non_terminal': A,
                                    'terminal': terminal,
                                    'production1': existing,
                                    'production2': prod_str,
                                    'type': 'FIRST-FOLLOW conflict'
                                })
                                self.parsing_table[A][terminal] = f"{existing} / {prod_str}"
                            else:
                                self.parsing_table[A][terminal] = prod_str
        
        print(f"Parsing table entries: {sum(len(row) for row in self.parsing_table.values())}")
        
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} parsing conflict(s)!")
        else:
            print("\n✓ No parsing conflicts found")
    
    def is_suitable_for_recursive_descent(self):
        """Check if grammar is suitable for Recursive Descent parsing"""
        # Main criteria: no left recursion
        return len(self.left_recursive_nts) == 0
    
    def is_backtrack_free(self):
        """Check if grammar can be parsed without backtracking (i.e., is LL(1))"""
        return (len(self.left_recursive_nts) == 0 and 
                len(self.left_factoring_needed) == 0 and 
                len(self.conflicts) == 0)
    
    def generate_parser_code(self, A):
        """Generate sample recursive descent parser code for a non-terminal"""
        code_lines = []
        code_lines.append(f"def parse_{A}():")
        code_lines.append(f"    \"\"\"Parse non-terminal {A}\"\"\"")
        
        productions = self.grammar.get(A, [])
        
        if len(productions) == 1:
            prod = productions[0]
            if prod == ['ε']:
                code_lines.append(f"    # ε production - do nothing")
                code_lines.append(f"    pass")
            else:
                for symbol in prod:
                    if symbol in self.non_terminals:
                        code_lines.append(f"    parse_{symbol}()")
                    else:
                        code_lines.append(f"    match('{symbol}')")
        else:
            # Multiple productions - need to look at current token
            code_lines.append(f"    if lookahead in {list(self.first.get(A, set()) - {'ε'})}:")
            
            for i, prod in enumerate(productions):
                first_set = self.compute_first_of_string(prod)
                
                if i == 0:
                    code_lines.append(f"        # {A} → {' '.join(prod)}")
                    code_lines.append(f"        if lookahead in {list(first_set - {'ε'})}:")
                else:
                    code_lines.append(f"        # {A} → {' '.join(prod)}")
                    code_lines.append(f"        elif lookahead in {list(first_set - {'ε'})}:")
                
                if prod == ['ε']:
                    code_lines.append(f"            pass  # ε production")
                else:
                    for symbol in prod:
                        if symbol in self.non_terminals:
                            code_lines.append(f"            parse_{symbol}()")
                        else:
                            code_lines.append(f"            match('{symbol}')")
            
            # Check FOLLOW set for epsilon productions
            if any('ε' in self.compute_first_of_string(p) for p in productions):
                code_lines.append(f"        elif lookahead in {list(self.follow.get(A, set()))}:")
                code_lines.append(f"            pass  # ε production via FOLLOW")
            
            code_lines.append(f"        else:")
            code_lines.append(f"            error('Unexpected token in {A}')")
        
        return '\n'.join(code_lines)
    
    def generate_excel(self, filename='RD_analysis.xlsx'):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_left_recursion_sheet(wb)
        self.create_left_factoring_sheet(wb)
        self.create_first_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_parsing_table_sheet(wb)
        self.create_parser_code_sheet(wb)
        self.create_recommendations_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "Recursive Descent Parser Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Suitable for Recursive Descent?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_suitable = self.is_suitable_for_recursive_descent()
        ws['B3'] = "YES ✓" if is_suitable else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_suitable else "FF0000")
        
        ws['A5'] = "Backtrack-Free (LL(1))?"
        ws['A5'].font = Font(bold=True, size=14)
        
        is_backtrack_free = self.is_backtrack_free()
        ws['B5'] = "YES ✓" if is_backtrack_free else "NO ✗"
        ws['B5'].font = Font(bold=True, size=14, color="008000" if is_backtrack_free else "FFA500")
        
        ws['A7'] = "Number of Non-Terminals:"
        ws['A7'].font = Font(bold=True)
        ws['B7'] = len(self.non_terminals)
        
        ws['A8'] = "Number of Terminals:"
        ws['A8'].font = Font(bold=True)
        ws['B8'] = len(self.terminals) - 1
        
        ws['A9'] = "Direct Left Recursions:"
        ws['A9'].font = Font(bold=True)
        ws['B9'] = len(self.direct_left_recursion)
        if len(self.direct_left_recursion) > 0:
            ws['B9'].font = Font(color="FF0000", bold=True)
        
        ws['A10'] = "Indirect Left Recursions:"
        ws['A10'].font = Font(bold=True)
        ws['B10'] = len(self.indirect_left_recursion)
        if len(self.indirect_left_recursion) > 0:
            ws['B10'].font = Font(color="FF0000", bold=True)
        
        ws['A11'] = "Left Factoring Needed:"
        ws['A11'].font = Font(bold=True)
        ws['B11'] = len(self.left_factoring_needed)
        if len(self.left_factoring_needed) > 0:
            ws['B11'].font = Font(color="FFA500", bold=True)
        
        ws['A12'] = "Parsing Conflicts:"
        ws['A12'].font = Font(bold=True)
        ws['B12'] = len(self.conflicts)
        if len(self.conflicts) > 0:
            ws['B12'].font = Font(color="FFA500", bold=True)
        
        # Summary
        ws['A14'] = "Summary:"
        ws['A14'].font = Font(bold=True, size=12)
        ws.merge_cells('A14:D14')
        
        row = 15
        if is_backtrack_free:
            ws[f'A{row}'] = "✓ Grammar is LL(1) - Perfect for backtrack-free Recursive Descent!"
            ws[f'A{row}'].font = Font(bold=True, color="008000")
            ws.merge_cells(f'A{row}:D{row}')
        elif is_suitable:
            ws[f'A{row}'] = "⚠ Grammar can use Recursive Descent but may need backtracking"
            ws[f'A{row}'].font = Font(bold=True, color="FFA500")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "Consider left factoring to eliminate backtracking"
            ws.merge_cells(f'A{row}:D{row}')
        else:
            ws[f'A{row}'] = "✗ Grammar has left recursion - NOT suitable for Recursive Descent"
            ws[f'A{row}'].font = Font(bold=True, color="FF0000")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "Left recursion must be eliminated first!"
            ws[f'A{row}'].font = Font(color="FF0000")
            ws.merge_cells(f'A{row}:D{row}')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Production #"
        ws['B1'] = "Non-Terminal"
        ws['C1'] = "Production"
        ws['D1'] = "Issues"
        
        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        prod_num = 0
        
        for nt in sorted(self.grammar.keys()):
            for prod in self.grammar[nt]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = nt
                ws[f'C{row}'] = f"{nt} → {' '.join(prod)}"
                
                # Check for issues
                issues = []
                if prod and prod[0] == nt:
                    issues.append("Direct Left Recursion")
                    ws[f'D{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws[f'D{row}'].font = Font(color="FFFFFF", bold=True)
                
                ws[f'D{row}'] = ', '.join(issues) if issues else 'OK'
                
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
    
    def create_left_recursion_sheet(self, wb):
        """Create left recursion analysis sheet"""
        ws = wb.create_sheet("Left Recursion")
        
        ws['A1'] = "Left Recursion Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Type"
        ws['B3'] = "Non-Terminal"
        ws['C3'] = "Production"
        ws['D3'] = "Description"
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=11, color="FFFFFF")
        
        row = 4
        
        if not self.direct_left_recursion and not self.indirect_left_recursion:
            ws['A4'] = "✓ No left recursion found - Grammar is suitable for Recursive Descent!"
            ws['A4'].font = Font(bold=True, color="008000")
            ws.merge_cells('A4:D4')
        else:
            for dlr in self.direct_left_recursion:
                ws[f'A{row}'] = "Direct"
                ws[f'B{row}'] = dlr['non_terminal']
                ws[f'C{row}'] = dlr['production']
                ws[f'D{row}'] = "Production starts with same non-terminal"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
            
            for ilr in self.indirect_left_recursion:
                ws[f'A{row}'] = "Indirect"
                ws[f'B{row}'] = ilr['cycle'][0]
                ws[f'C{row}'] = ilr['description']
                ws[f'D{row}'] = "Cycle in derivation graph"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
    
    def create_left_factoring_sheet(self, wb):
        """Create left factoring analysis sheet"""
        ws = wb.create_sheet("Left Factoring")
        
        ws['A1'] = "Left Factoring Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Non-Terminal"
        ws['B3'] = "Common Prefix"
        ws['C3'] = "Productions"
        ws['D3'] = "Impact"
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            ws[cell].font = Font(bold=True, size=11, color="FFFFFF")
        
        row = 4
        
        if not self.left_factoring_needed:
            ws['A4'] = "✓ No left factoring needed - Grammar is left-factored!"
            ws['A4'].font = Font(bold=True, color="008000")
            ws.merge_cells('A4:D4')
        else:
            for lf in self.left_factoring_needed:
                ws[f'A{row}'] = lf['non_terminal']
                ws[f'B{row}'] = ' '.join(lf['common_prefix'])
                ws[f'C{row}'] = '\n'.join(lf['productions'])
                ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws[f'D{row}'] = "May require backtracking"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                
                ws.row_dimensions[row].height = 15 * len(lf['productions'])
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
    
    def create_first_sheet(self, wb):
        """Create FIRST sets sheet"""
        ws = wb.create_sheet("FIRST Sets")
        
        ws['A1'] = "Symbol"
        ws['B1'] = "FIRST Set"
        ws['C1'] = "Used for Parsing Decisions"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            first_set = sorted(self.first.get(A, set()))
            ws[f'B{row}'] = ', '.join(first_set)
            
            if 'ε' in self.first.get(A, set()):
                ws[f'C{row}'] = "Uses FOLLOW set"
            else:
                ws[f'C{row}'] = "Direct lookahead"
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
    
    def create_follow_sheet(self, wb):
        """Create FOLLOW sets sheet"""
        ws = wb.create_sheet("FOLLOW Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        ws['C1'] = "Used for ε-productions"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            follow_set = sorted(self.follow.get(A, set()))
            ws[f'B{row}'] = ', '.join(follow_set)
            
            has_epsilon = any('ε' in self.compute_first_of_string(prod) 
                            for prod in self.grammar.get(A, []))
            ws[f'C{row}'] = "Yes" if has_epsilon else "No"
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
    
    def create_parsing_table_sheet(self, wb):
        """Create parsing table sheet"""
        ws = wb.create_sheet("Parsing Table")
        
        ws['A1'] = "Predictive Parsing Table"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:' + get_column_letter(len(self.terminals) + 1) + '1')
        
        # Headers
        ws['A3'] = "Non-Terminal"
        ws['A3'].font = Font(bold=True, size=11)
        ws['A3'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws['A3'].font = Font(bold=True, size=11, color="FFFFFF")
        
        col_idx = 2
        terminal_cols = {}
        for terminal in sorted(self.terminals):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}3'] = terminal
            ws[f'{col_letter}3'].font = Font(bold=True, size=11)
            ws[f'{col_letter}3'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[f'{col_letter}3'].font = Font(bold=True, size=11, color="FFFFFF")
            terminal_cols[terminal] = col_letter
            col_idx += 1
        
        # Fill table
        row = 4
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'A{row}'].font = Font(bold=True)
            
            for terminal in sorted(self.terminals):
                col_letter = terminal_cols[terminal]
                entry = self.parsing_table.get(A, {}).get(terminal, '')
                ws[f'{col_letter}{row}'] = entry
                
                # Highlight conflicts
                if '/' in str(entry):
                    ws[f'{col_letter}{row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    ws[f'{col_letter}{row}'].font = Font(color="FFFFFF", bold=True)
                elif entry:
                    ws[f'{col_letter}{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        for col_letter in terminal_cols.values():
            ws.column_dimensions[col_letter].width = 30
    
    def create_parser_code_sheet(self, wb):
        """Create parser code sheet"""
        ws = wb.create_sheet("Parser Code")
        
        ws['A1'] = "Generated Recursive Descent Parser Code"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "This is sample code for a Recursive Descent parser."
        ws['A3'].font = Font(italic=True)
        ws.merge_cells('A3:B3')
        
        ws['A4'] = "Note: Adjust according to your specific grammar and requirements."
        ws['A4'].font = Font(italic=True)
        ws.merge_cells('A4:B4')
        
        row = 6
        
        # Header code
        ws[f'A{row}'] = "# Recursive Descent Parser"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        ws[f'A{row}'] = ""
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Helper functions
        helper_code = """# Global variables
lookahead = None
tokens = []
index = 0

def match(expected):
    \"\"\"Match expected terminal and advance\"\"\"
    global lookahead, index
    if lookahead == expected:
        index += 1
        if index < len(tokens):
            lookahead = tokens[index]
        else:
            lookahead = '$'
    else:
        error(f'Expected {expected}, got {lookahead}')

def error(msg):
    \"\"\"Report parsing error\"\"\"
    raise Exception(f'Parse error: {msg}')

def parse(input_tokens):
    \"\"\"Main parsing function\"\"\"
    global lookahead, tokens, index
    tokens = input_tokens + ['$']
    index = 0
    lookahead = tokens[0]
    parse_""" + (self.start_symbol if self.start_symbol else "S") + """()
    if lookahead != '$':
        error('Unexpected tokens after parsing')
"""
        
        for line in helper_code.split('\n'):
            ws[f'A{row}'] = line
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        ws[f'A{row}'] = ""
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Generate parser functions for each non-terminal
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = f"# Parser for {A}"
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            parser_code = self.generate_parser_code(A)
            for line in parser_code.split('\n'):
                ws[f'A{row}'] = line
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
            
            ws[f'A{row}'] = ""
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 20
    
    def create_recommendations_sheet(self, wb):
        """Create recommendations sheet"""
        ws = wb.create_sheet("Recommendations")
        
        ws['A1'] = "Recommendations for Recursive Descent Parsing"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Check current status
        has_left_recursion = len(self.left_recursive_nts) > 0
        needs_left_factoring = len(self.left_factoring_needed) > 0
        has_conflicts = len(self.conflicts) > 0
        
        if not has_left_recursion and not needs_left_factoring and not has_conflicts:
            ws[f'A{row}'] = "✓ EXCELLENT!"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="008000")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws[f'A{row}'] = "Your grammar is LL(1) and perfect for Recursive Descent parsing!"
            ws[f'A{row}'].font = Font(size=12, color="008000")
            ws.merge_cells(f'A{row}:C{row}')
            row += 2
            
            ws[f'A{row}'] = "Implementation Steps:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            steps = [
                "1. Use the parsing table to guide parsing decisions",
                "2. Implement one function per non-terminal",
                "3. Use lookahead token to select production",
                "4. No backtracking needed - deterministic parsing"
            ]
            
            for step in steps:
                ws[f'A{row}'] = step
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
        else:
            ws[f'A{row}'] = "⚠ GRAMMAR NEEDS TRANSFORMATION"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="FF0000" if has_left_recursion else "FFA500")
            ws.merge_cells(f'A{row}:C{row}')
            row += 2
            
            if has_left_recursion:
                ws[f'A{row}'] = "CRITICAL: Eliminate Left Recursion"
                ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                ws[f'A{row}'] = "Left recursion prevents Recursive Descent parsing from working."
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                ws[f'A{row}'] = "Affected non-terminals: " + ', '.join(self.left_recursive_nts)
                ws[f'A{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                ws.merge_cells(f'A{row}:C{row}')
                row += 2
                
                ws[f'A{row}'] = "Algorithm to eliminate left recursion:"
                ws[f'A{row}'].font = Font(bold=True)
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                elimination_steps = [
                    "1. For direct left recursion A → Aα | β:",
                    "   Replace with: A → βA' and A' → αA' | ε",
                    "",
                    "2. For indirect left recursion:",
                    "   a. Order non-terminals A1, A2, ..., An",
                    "   b. For each Ai, eliminate all Ai → Aj... where j < i",
                    "   c. Eliminate direct left recursion for Ai"
                ]
                
                for step in elimination_steps:
                    ws[f'A{row}'] = step
                    ws.merge_cells(f'A{row}:C{row}')
                    row += 1
                
                row += 1
            
            if needs_left_factoring:
                ws[f'A{row}'] = "RECOMMENDED: Apply Left Factoring"
                ws[f'A{row}'].font = Font(bold=True, size=12, color="FFA500")
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                ws[f'A{row}'] = "Left factoring eliminates backtracking and makes the parser more efficient."
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                ws[f'A{row}'] = f"Found {len(self.left_factoring_needed)} case(s) that need left factoring."
                ws[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                ws.merge_cells(f'A{row}:C{row}')
                row += 2
                
                ws[f'A{row}'] = "Algorithm for left factoring:"
                ws[f'A{row}'].font = Font(bold=True)
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                factoring_steps = [
                    "1. Find productions with common prefix: A → αβ1 | αβ2 | ... | αβn | γ",
                    "2. Replace with: A → αA' | γ and A' → β1 | β2 | ... | βn",
                    "3. Repeat until no common prefixes remain"
                ]
                
                for step in factoring_steps:
                    ws[f'A{row}'] = step
                    ws.merge_cells(f'A{row}:C{row}')
                    row += 1
                
                row += 1
            
            if has_conflicts:
                ws[f'A{row}'] = "WARNING: Parsing Conflicts Detected"
                ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                ws[f'A{row}'] = f"Found {len(self.conflicts)} conflict(s) in the parsing table."
                ws.merge_cells(f'A{row}:C{row}')
                row += 1
                
                ws[f'A{row}'] = "This indicates the grammar is not LL(1)."
                ws[f'A{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                ws.merge_cells(f'A{row}:C{row}')
                row += 2
        
        row += 1
        ws[f'A{row}'] = "Additional Resources:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        resources = [
            "• Check the 'Left Recursion' sheet for specific problematic productions",
            "• Check the 'Left Factoring' sheet for common prefix conflicts",
            "• Use the 'Parsing Table' sheet to identify conflicts",
            "• Refer to the 'Parser Code' sheet for implementation examples"
        ]
        
        for resource in resources:
            ws[f'A{row}'] = resource
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    def analyze(self):
        """Run complete analysis"""
        print("=" * 70)
        print("RECURSIVE DESCENT PARSER ANALYZER")
        print("=" * 70)
        
        self.check_direct_left_recursion()
        self.check_indirect_left_recursion()
        self.check_left_factoring()
        self.compute_first()
        self.compute_follow()
        self.build_parsing_table()
        
        print("\n" + "=" * 70)
        print("ANALYSIS RESULT")
        print("=" * 70)
        
        is_suitable = self.is_suitable_for_recursive_descent()
        is_ll1 = self.is_backtrack_free()
        
        if is_ll1:
            print("\n✅ Grammar is LL(1) - Perfect for Recursive Descent!")
            print("   - No left recursion")
            print("   - No left factoring needed")
            print("   - No parsing conflicts")
            print("   - Deterministic, backtrack-free parsing possible")
        elif is_suitable:
            print("\n⚠️  Grammar is suitable for Recursive Descent but NOT LL(1)")
            print("   - No left recursion ✓")
            if self.left_factoring_needed:
                print(f"   - Left factoring needed ({len(self.left_factoring_needed)} cases)")
            if self.conflicts:
                print(f"   - Parsing conflicts exist ({len(self.conflicts)} conflicts)")
            print("   - May require backtracking")
        else:
            print("\n❌ Grammar is NOT suitable for Recursive Descent")
            print(f"   - Has left recursion: {', '.join(self.left_recursive_nts)}")
            print("   - Must eliminate left recursion first!")
        
        print("\n" + "=" * 70)


def main():
    """Main function"""
    print("\n" + "=" * 70)
    print("RECURSIVE DESCENT PARSER ANALYZER")
    print("=" * 70)
    print("\nEnter your grammar (use → for production arrow)")
    print("Format: E → E + T | T")
    print("Use | for alternatives")
    print("Use ε or epsilon for empty production")
    print("Enter a blank line when done:")
    print()
    
    
    
    
    # Analyze grammar
    analyzer = RecursiveDescentAnalyzer(grammar)
    analyzer.analyze()
    
    analyzer.generate_excel()
    
    print("\n" + "=" * 70)
    print("Analysis complete! Check the Excel file for detailed results.")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()