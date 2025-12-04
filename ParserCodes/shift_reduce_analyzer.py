"""
Shift-Reduce Conflict Analyzer
Detects shift-reduce and reduce-reduce conflicts in grammars
Uses LR(0) automaton to identify potential parsing conflicts
"""

from collections import defaultdict, deque
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class LR0Item:
    """Represents an LR(0) item: [A → α·β]"""
    def __init__(self, lhs, rhs, dot_pos):
        self.lhs = lhs
        self.rhs = rhs
        self.dot_pos = dot_pos
    
    def __eq__(self, other):
        return (self.lhs == other.lhs and 
                self.rhs == other.rhs and 
                self.dot_pos == other.dot_pos)
    
    def __hash__(self):
        return hash((self.lhs, tuple(self.rhs), self.dot_pos))
    
    def __repr__(self):
        rhs_with_dot = self.rhs[:self.dot_pos] + ['·'] + self.rhs[self.dot_pos:]
        return f"[{self.lhs} → {' '.join(rhs_with_dot)}]"
    
    def is_complete(self):
        return self.dot_pos >= len(self.rhs)
    
    def next_symbol(self):
        if self.is_complete():
            return None
        return self.rhs[self.dot_pos]
    
    def advance(self):
        return LR0Item(self.lhs, self.rhs, self.dot_pos + 1)


class LR0State:
    """Represents a state in LR(0) automaton"""
    def __init__(self, state_id, items):
        self.id = state_id
        self.items = frozenset(items)
    
    def __eq__(self, other):
        return self.items == other.items
    
    def __hash__(self):
        return hash(self.items)
    
    def __repr__(self):
        return f"State {self.id}:\n" + "\n".join(f"  {item}" for item in self.items)


class ShiftReduceAnalyzer:
    def __init__(self, grammar_text):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start_symbol = None
        self.augmented_start = None
        
        self.states = []
        self.state_map = {}
        self.transitions = {}
        
        self.shift_reduce_conflicts = []
        self.reduce_reduce_conflicts = []
        self.conflict_states = set()
        
        self.parse_grammar(grammar_text)
        self.augment_grammar()
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if self.start_symbol is None:
                    self.start_symbol = lhs
                
                if lhs not in self.grammar:
                    self.grammar[lhs] = []
                self.grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol not in self.non_terminals:
                        self.terminals.add(symbol)
        
        self.terminals.add('$')
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols
    
    def augment_grammar(self):
        """Augment grammar with SBar → S"""
        self.augmented_start = self.start_symbol + "Bar"
        self.grammar[self.augmented_start] = [[self.start_symbol]]
        self.non_terminals.add(self.augmented_start)
    
    def closure(self, items):
        """Compute closure of a set of LR(0) items"""
        closure_set = set(items)
        changed = True
        
        while changed:
            changed = False
            new_items = set()
            
            for item in closure_set:
                next_sym = item.next_symbol()
                if next_sym and next_sym in self.non_terminals:
                    # Add items for all productions of next_sym
                    for production in self.grammar.get(next_sym, []):
                        new_item = LR0Item(next_sym, production, 0)
                        if new_item not in closure_set:
                            new_items.add(new_item)
                            changed = True
            
            closure_set.update(new_items)
        
        return closure_set
    
    def goto(self, items, symbol):
        """Compute GOTO(items, symbol) for LR(0)"""
        moved_items = set()
        
        for item in items:
            if item.next_symbol() == symbol:
                moved_items.add(item.advance())
        
        if not moved_items:
            return None
        
        return self.closure(moved_items)
    
    def build_automaton(self):
        """Build LR(0) automaton"""
        print("\n=== BUILDING LR(0) AUTOMATON ===\n")
        
        # Create initial item
        initial_item = LR0Item(self.augmented_start, 
                              self.grammar[self.augmented_start][0], 
                              0)
        initial_items = self.closure({initial_item})
        initial_items_frozen = frozenset(initial_items)
        
        initial_state = LR0State(0, initial_items)
        self.states.append(initial_state)
        self.state_map[initial_items_frozen] = initial_state
        
        queue = deque([initial_state])
        
        while queue:
            current_state = queue.popleft()
            
            # Find all symbols that can be shifted
            symbols_to_process = set()
            for item in current_state.items:
                next_sym = item.next_symbol()
                if next_sym:
                    symbols_to_process.add(next_sym)
            
            for symbol in symbols_to_process:
                goto_items = self.goto(current_state.items, symbol)
                
                if goto_items:
                    goto_items_frozen = frozenset(goto_items)
                    
                    if goto_items_frozen in self.state_map:
                        next_state = self.state_map[goto_items_frozen]
                    else:
                        next_state = LR0State(len(self.states), goto_items)
                        self.states.append(next_state)
                        self.state_map[goto_items_frozen] = next_state
                        queue.append(next_state)
                    
                    self.transitions[(current_state.id, symbol)] = next_state.id
        
        print(f"Created {len(self.states)} states")
        print(f"Created {len(self.transitions)} transitions")
    
    def analyze_conflicts(self):
        """Analyze states for shift-reduce and reduce-reduce conflicts"""
        print("\n=== ANALYZING SHIFT-REDUCE CONFLICTS ===\n")
        
        for state in self.states:
            # Separate shift items and reduce items
            shift_items = []
            reduce_items = []
            
            for item in state.items:
                if item.is_complete():
                    # Skip augmented start production
                    if item.lhs != self.augmented_start:
                        reduce_items.append(item)
                else:
                    next_sym = item.next_symbol()
                    if next_sym in self.terminals:
                        shift_items.append(item)
            
            # Check for shift-reduce conflicts
            if shift_items and reduce_items:
                self.conflict_states.add(state.id)
                
                # Get shift symbols
                shift_symbols = set()
                for item in shift_items:
                    shift_symbols.add(item.next_symbol())
                
                for reduce_item in reduce_items:
                    for shift_symbol in shift_symbols:
                        prod_num = self.get_production_number(reduce_item.lhs, reduce_item.rhs)
                        
                        self.shift_reduce_conflicts.append({
                            'state': state.id,
                            'symbol': shift_symbol,
                            'shift_item': str([item for item in shift_items if item.next_symbol() == shift_symbol][0]),
                            'reduce_item': str(reduce_item),
                            'production': prod_num
                        })
            
            # Check for reduce-reduce conflicts
            if len(reduce_items) > 1:
                self.conflict_states.add(state.id)
                
                for i in range(len(reduce_items)):
                    for j in range(i + 1, len(reduce_items)):
                        prod_num1 = self.get_production_number(reduce_items[i].lhs, reduce_items[i].rhs)
                        prod_num2 = self.get_production_number(reduce_items[j].lhs, reduce_items[j].rhs)
                        
                        self.reduce_reduce_conflicts.append({
                            'state': state.id,
                            'reduce_item1': str(reduce_items[i]),
                            'reduce_item2': str(reduce_items[j]),
                            'production1': prod_num1,
                            'production2': prod_num2
                        })
        
        total_conflicts = len(self.shift_reduce_conflicts) + len(self.reduce_reduce_conflicts)
        
        print(f"Found {len(self.shift_reduce_conflicts)} shift-reduce conflict(s)")
        print(f"Found {len(self.reduce_reduce_conflicts)} reduce-reduce conflict(s)")
        print(f"Total conflicts: {total_conflicts}")
        print(f"Conflict states: {len(self.conflict_states)}")
        
        if total_conflicts == 0:
            print("\n✓ No conflicts found - Grammar is conflict-free!")
        else:
            print(f"\n⚠️  Grammar has conflicts in {len(self.conflict_states)} state(s)!")
    
    def get_production_number(self, lhs, rhs):
        """Get production number for A → α"""
        prod_num = 0
        
        # First, check augmented production (always production 0)
        if lhs == self.augmented_start:
            for prod in self.grammar[self.augmented_start]:
                if prod == rhs:
                    return prod_num
                prod_num += 1
        
        # Then number the rest in order (excluding augmented start)
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def has_conflicts(self):
        """Check if grammar has any conflicts"""
        return len(self.shift_reduce_conflicts) + len(self.reduce_reduce_conflicts) > 0
    
    def generate_excel(self, filename='shift_reduce_analysis.xlsx'):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_shift_reduce_conflicts_sheet(wb)
        self.create_reduce_reduce_conflicts_sheet(wb)
        self.create_states_sheet(wb)
        self.create_conflict_states_sheet(wb)
        self.create_transitions_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "Shift-Reduce Conflict Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Conflict-Free?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_conflict_free = not self.has_conflicts()
        ws['B3'] = "YES ✓" if is_conflict_free else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_conflict_free else "FF0000")
        
        ws['A5'] = "Number of States:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.states)
        
        ws['A6'] = "States with Conflicts:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = len(self.conflict_states)
        ws['B6'].font = Font(color="FF0000" if len(self.conflict_states) > 0 else "008000")
        
        ws['A8'] = "Shift-Reduce Conflicts:"
        ws['A8'].font = Font(bold=True)
        ws['B8'] = len(self.shift_reduce_conflicts)
        ws['B8'].font = Font(color="FF0000" if len(self.shift_reduce_conflicts) > 0 else "008000")
        
        ws['A9'] = "Reduce-Reduce Conflicts:"
        ws['A9'].font = Font(bold=True)
        ws['B9'] = len(self.reduce_reduce_conflicts)
        ws['B9'].font = Font(color="FF0000" if len(self.reduce_reduce_conflicts) > 0 else "008000")
        
        ws['A10'] = "Total Conflicts:"
        ws['A10'].font = Font(bold=True)
        total_conflicts = len(self.shift_reduce_conflicts) + len(self.reduce_reduce_conflicts)
        ws['B10'] = total_conflicts
        ws['B10'].font = Font(bold=True, color="FF0000" if total_conflicts > 0 else "008000")
        
        if not is_conflict_free:
            ws['A12'] = "⚠️ Grammar has parsing conflicts!"
            ws['A12'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A12:D12')
            
            ws['A13'] = "See conflict details in separate sheets."
            ws['A13'].font = Font(size=11, color="FF0000")
            ws.merge_cells('A13:D13')
        else:
            ws['A12'] = "✓ Grammar is conflict-free!"
            ws['A12'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A12:D12')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Production #"
        ws['B1'] = "Non-Terminal"
        ws['C1'] = "Production"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        prod_num = 0
        
        # First, add augmented production (production 0)
        for prod in self.grammar[self.augmented_start]:
            ws[f'A{row}'] = prod_num
            ws[f'B{row}'] = self.augmented_start
            ws[f'C{row}'] = f"{self.augmented_start} → {' '.join(prod)}"
            ws[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'C{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            row += 1
            prod_num += 1
        
        # Then add the rest in sorted order
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = nt
                ws[f'C{row}'] = f"{nt} → {' '.join(prod)}"
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def create_shift_reduce_conflicts_sheet(self, wb):
        """Create shift-reduce conflicts sheet"""
        ws = wb.create_sheet("Shift-Reduce Conflicts")
        
        ws['A1'] = "State"
        ws['B1'] = "Symbol"
        ws['C1'] = "Shift Item"
        ws['D1'] = "Reduce Item"
        ws['E1'] = "Production #"
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            ws[cell].font = Font(bold=True, size=11, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        if self.shift_reduce_conflicts:
            row = 2
            for conflict in self.shift_reduce_conflicts:
                ws[f'A{row}'] = conflict['state']
                ws[f'B{row}'] = conflict['symbol']
                ws[f'C{row}'] = conflict['shift_item']
                ws[f'D{row}'] = conflict['reduce_item']
                ws[f'E{row}'] = conflict['production']
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        else:
            ws['A2'] = "No shift-reduce conflicts found"
            ws['A2'].font = Font(size=11, color="008000")
            ws.merge_cells('A2:E2')
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
    
    def create_reduce_reduce_conflicts_sheet(self, wb):
        """Create reduce-reduce conflicts sheet"""
        ws = wb.create_sheet("Reduce-Reduce Conflicts")
        
        ws['A1'] = "State"
        ws['B1'] = "Reduce Item 1"
        ws['C1'] = "Production #1"
        ws['D1'] = "Reduce Item 2"
        ws['E1'] = "Production #2"
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=11, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        if self.reduce_reduce_conflicts:
            row = 2
            for conflict in self.reduce_reduce_conflicts:
                ws[f'A{row}'] = conflict['state']
                ws[f'B{row}'] = conflict['reduce_item1']
                ws[f'C{row}'] = conflict['production1']
                ws[f'D{row}'] = conflict['reduce_item2']
                ws[f'E{row}'] = conflict['production2']
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                
                row += 1
        else:
            ws['A2'] = "No reduce-reduce conflicts found"
            ws['A2'].font = Font(size=11, color="008000")
            ws.merge_cells('A2:E2')
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
    
    def create_states_sheet(self, wb):
        """Create states sheet"""
        ws = wb.create_sheet("States")
        
        ws['A1'] = "State"
        ws['B1'] = "LR(0) Items"
        ws['C1'] = "Conflict?"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for state in self.states:
            ws[f'A{row}'] = f"State {state.id}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_text = "\n".join(str(item) for item in sorted(state.items, key=str))
            ws[f'B{row}'] = items_text
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            if state.id in self.conflict_states:
                ws[f'C{row}'] = "YES"
                ws[f'C{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f'C{row}'].font = Font(color="FFFFFF", bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            else:
                ws[f'C{row}'] = "No"
            
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[row].height = 15 * len(state.items)
            
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
    
    def create_conflict_states_sheet(self, wb):
        """Create conflict states detail sheet"""
        ws = wb.create_sheet("Conflict States Detail")
        
        ws['A1'] = "State"
        ws['B1'] = "Items"
        ws['C1'] = "Conflict Type"
        ws['D1'] = "Description"
        
        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        if self.conflict_states:
            row = 2
            for state_id in sorted(self.conflict_states):
                state = self.states[state_id]
                
                ws[f'A{row}'] = state_id
                ws[f'A{row}'].font = Font(bold=True)
                
                items_text = "\n".join(str(item) for item in sorted(state.items, key=str))
                ws[f'B{row}'] = items_text
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                
                # Determine conflict types in this state
                conflict_types = []
                if any(c['state'] == state_id for c in self.shift_reduce_conflicts):
                    conflict_types.append("Shift-Reduce")
                if any(c['state'] == state_id for c in self.reduce_reduce_conflicts):
                    conflict_types.append("Reduce-Reduce")
                
                ws[f'C{row}'] = ", ".join(conflict_types)
                ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='center')
                
                # Count conflicts
                sr_count = sum(1 for c in self.shift_reduce_conflicts if c['state'] == state_id)
                rr_count = sum(1 for c in self.reduce_reduce_conflicts if c['state'] == state_id)
                
                ws[f'D{row}'] = f"SR: {sr_count}, RR: {rr_count}"
                ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                ws.row_dimensions[row].height = 15 * len(state.items)
                
                row += 1
        else:
            ws['A2'] = "No conflict states found"
            ws['A2'].font = Font(size=11, color="008000")
            ws.merge_cells('A2:D2')
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def create_transitions_sheet(self, wb):
        """Create transitions sheet"""
        ws = wb.create_sheet("Transitions")
        
        ws['A1'] = "From State"
        ws['B1'] = "Symbol"
        ws['C1'] = "To State"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for (from_state, symbol), to_state in sorted(self.transitions.items()):
            ws[f'A{row}'] = from_state
            ws[f'B{row}'] = symbol
            ws[f'C{row}'] = to_state
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15


def main():
    from grammar import grammar
    
    print("="*70)
    print("SHIFT-REDUCE CONFLICT ANALYZER")
    print("="*70)
    
    analyzer = ShiftReduceAnalyzer(grammar)
    
    print(f"\nGrammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)} terminals")
    print(f"Start symbol: {analyzer.start_symbol}")
    print(f"Augmented start: {analyzer.augmented_start}")
    
    analyzer.build_automaton()
    analyzer.analyze_conflicts()
    analyzer.generate_excel()
    
    print("\n" + "="*70)
    print("FINAL RESULT")
    print("="*70)
    
    if not analyzer.has_conflicts():
        print("✅ The grammar is CONFLICT-FREE")
        print("\nNo shift-reduce or reduce-reduce conflicts detected!")
    else:
        print("❌ The grammar HAS CONFLICTS")
        print(f"\nShift-Reduce Conflicts: {len(analyzer.shift_reduce_conflicts)}")
        print(f"Reduce-Reduce Conflicts: {len(analyzer.reduce_reduce_conflicts)}")
        print(f"Total Conflicts: {len(analyzer.shift_reduce_conflicts) + len(analyzer.reduce_reduce_conflicts)}")
        print(f"\nConflict States: {sorted(analyzer.conflict_states)}")
        
        if analyzer.shift_reduce_conflicts:
            print(f"\n--- Shift-Reduce Conflicts ---")
            for i, conflict in enumerate(analyzer.shift_reduce_conflicts[:5], 1):  # Show first 5
                print(f"\n  Conflict {i}:")
                print(f"    State: {conflict['state']}")
                print(f"    Symbol: {conflict['symbol']}")
                print(f"    Shift: {conflict['shift_item']}")
                print(f"    Reduce: {conflict['reduce_item']}")
            if len(analyzer.shift_reduce_conflicts) > 5:
                print(f"\n  ... and {len(analyzer.shift_reduce_conflicts) - 5} more (see Excel file)")
        
        if analyzer.reduce_reduce_conflicts:
            print(f"\n--- Reduce-Reduce Conflicts ---")
            for i, conflict in enumerate(analyzer.reduce_reduce_conflicts[:5], 1):  # Show first 5
                print(f"\n  Conflict {i}:")
                print(f"    State: {conflict['state']}")
                print(f"    Reduce 1: {conflict['reduce_item1']}")
                print(f"    Reduce 2: {conflict['reduce_item2']}")
            if len(analyzer.reduce_reduce_conflicts) > 5:
                print(f"\n  ... and {len(analyzer.reduce_reduce_conflicts) - 5} more (see Excel file)")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()
